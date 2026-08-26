from __future__ import annotations
from apps.imports.utils import get_tech_stock_location
from apps.orders.models import OrderItem, Order
from apps.stock.models import Product
import os
import re
from typing import Any, Dict, List, Tuple, Optional

from django.core.exceptions import ValidationError, FieldDoesNotExist
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from apps.clients.models import Client
from apps.imports.models import RawExcelRow, ImportBatch
from apps.imports.rules import EXACT, CONTAINS, PRODUCT_KEYWORDS




def _s(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def normalize_header(text: str) -> str:
    if not text:
        return ""
    text = str(text).lower().strip()

    text = re.sub(r"[^0-9a-zР°-СЏС‘]+", " ", text, flags=re.IGNORECASE)

    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalize_cell_value(val: Any) -> Any:
    """
    Р”РµР»Р°РµС‚ Р·РЅР°С‡РµРЅРёРµ Р±РµР·РѕРїР°СЃРЅС‹Рј РґР»СЏ JSONField.
    datetime/date -> ISO СЃС‚СЂРѕРєР°
    РѕСЃС‚Р°Р»СЊРЅРѕРµ -> РєР°Рє РµСЃС‚СЊ (str/int/float/bool/None)
    """
    if val is None:
        return None
    # openpyxl РјРѕР¶РµС‚ РѕС‚РґР°РІР°С‚СЊ datetime/date
    try:
        import datetime as _dt
        if isinstance(val, (_dt.datetime, _dt.date)):
            return val.isoformat()
    except Exception:
        pass
    return val


def read_xlsx(file_path: str) -> Tuple[List[str], List[Tuple[int, Tuple[Any, ...]]]]:
    wb = load_workbook(file_path)
    sheet = wb.active

    headers = [_s(c.value) for c in sheet[1]]
    rows: List[Tuple[int, Tuple[Any, ...]]] = []

    for row_num, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        rows.append((row_num, tuple(values)))

    return headers, rows


def row_to_dict(headers: List[str], values: Tuple[Any, ...]) -> Dict[str, Any]:
    data: Dict[str, Any] = {}
    for i, h in enumerate(headers):
        key = normalize_header(h)
        if not key:
            continue
        v = values[i] if i < len(values) else None
        data[key] = normalize_cell_value(v)
    return data


def pick_any(row: dict, exact=None, contains=None):
    exact = [normalize_header(x) for x in (exact or [])]
    contains = [normalize_header(x) for x in (contains or [])]

    for key, value in row.items():
        normalized = normalize_header(key)

        if normalized in exact:
            return value

        if any(c in normalized for c in contains):
            return value

    return None

def model_has_field(model, field_name: str) -> bool:
    try:
        model._meta.get_field(field_name)
        return True
    except FieldDoesNotExist:
        return False

def _is_empty_row(values) -> bool:
    def _empty(v):
        if v is None:
            return True
        if isinstance(v, str) and not v.strip():
            return True
        return False
    return all(_empty(v) for v in values)

def normalize_phone(val: Any) -> str:

    raw = _s(val)
    if not raw:
        return ""

    # Excel-С‡РёСЃР»Р° С‚РёРїР° 79161234567.0
    if raw.endswith(".0"):
        raw = raw[:-2]

    digits = re.sub(r"\D+", "", raw)


    if len(digits) == 10:
        digits = "7" + digits


    if len(digits) == 11 and digits[0] in ("7", "8"):
        if digits[0] == "8":
            digits = "7" + digits[1:]
        return digits

    return ""


def normalize_date(val: Any) -> str:
    """
    РџСЂРµРѕР±СЂР°Р·СѓРµС‚ РґР°С‚Сѓ РІ YYYY-MM-DD
    РџСЂРёРЅРёРјР°РµС‚: datetime, date, ISO СЃС‚СЂРѕРєСѓ, СЃС‚СЂРѕРєСѓ DD.MM.YYYY
    """
    if val is None:
        return ""

    # Р•СЃР»Рё СЌС‚Рѕ datetime/date РѕР±СЉРµРєС‚
    if hasattr(val, 'strftime'):
        return val.strftime('%Y-%m-%d')

    raw = _s(val)
    if not raw:
        return ""

    # Р•СЃР»Рё СЌС‚Рѕ ISO СЃС‚СЂРѕРєР° СЃ РІСЂРµРјРµРЅРµРј (1996-05-27T00:00:00)
    if 'T' in raw:
        return raw.split('T')[0]


    if re.match(r'^\d{4}-\d{2}-\d{2}$', raw):
        return raw


    match = re.search(r'^(\d{1,2})\.(\d{1,2})\.(\d{4})$', raw.strip())
    if match:
        day, month, year = match.groups()
        return f"{year}-{int(month):02d}-{int(day):02d}"

    return ""




def build_full_name(row: Dict[str, Any]) -> str:
    last_name = _s(pick_any(row, exact=EXACT["last_name"]))
    first_name = _s(pick_any(row, exact=EXACT["first_name"]))
    middle_name = _s(pick_any(row, exact=EXACT["middle_name"]))

    parts = [p for p in [last_name, first_name, middle_name] if p]
    return " ".join(parts)


def build_notes(row: Dict[str, Any]) -> str:
    """РЎРѕР±РёСЂР°РµС‚ РІСЃРµ РґРѕРїРѕР»РЅРёС‚РµР»СЊРЅС‹Рµ РїРѕР»СЏ РІ РѕРґРЅСѓ СЃС‚СЂРѕРєСѓ РґР»СЏ notes"""
    parts = []

    # РџСЂРѕС„РµСЃСЃРёСЏ
    profession = _s(pick_any(row, exact=EXACT["profession"]))
    if profession:
        parts.append(f"РџСЂРѕС„РµСЃСЃРёСЏ: {profession}")

    # РљР»СѓР±
    club = _s(pick_any(row, exact=EXACT["club"]))
    if club:
        parts.append(f"РљР»СѓР±: {club}")



    return " | ".join(parts)


def build_address(row: Dict[str, Any]) -> str:
    """РЎРѕР±РёСЂР°РµС‚ РїРѕР»РЅС‹Р№ Р°РґСЂРµСЃ РёР· СЃРѕСЃС‚Р°РІРЅС‹С… С‡Р°СЃС‚РµР№"""
    parts = []


    city = _s(pick_any(row, exact=EXACT["city"]))
    if city:
        parts.append(f"Рі. {city}")


    street = _s(pick_any(row, exact=EXACT["street"]))
    if street:
        parts.append(f"СѓР». {street}")

    # Р”РѕРј
    house = _s(pick_any(row, exact=EXACT["house"]))
    if house:
        parts.append(f"Рґ. {house}")

    # РљРІР°СЂС‚РёСЂР°
    flat = _s(pick_any(row, exact=EXACT["flat"]))
    if flat:
        parts.append(f"РєРІ. {flat}")

    # Р•СЃР»Рё РµСЃС‚СЊ РіРѕС‚РѕРІС‹Р№ Р°РґСЂРµСЃ (РЅР°РїСЂРёРјРµСЂ, РёР· РїРѕР»СЏ "РЈРєР°Р¶РёС‚Рµ Р°РґСЂРµСЃ РґР»СЏ РґРѕСЃС‚Р°РІРєРё")
    full_address = _s(pick_any(row, contains=CONTAINS.get("delivery_address_text", [])))
    if full_address and not parts:  # РёСЃРїРѕР»СЊР·СѓРµРј С‚РѕР»СЊРєРѕ РµСЃР»Рё РЅРµС‚ СЃРѕСЃС‚Р°РІРЅС‹С… С‡Р°СЃС‚РµР№
        return full_address

    return ", ".join(parts)

def is_empty_row(values) -> bool:
    """
    РџСЂРѕРІРµСЂСЏРµРј, РїРѕР»РЅРѕСЃС‚СЊСЋ Р»Рё СЃС‚СЂРѕРєР° РїСѓСЃС‚Р°СЏ.
    РџСѓСЃС‚Р°СЏ = РІСЃРµ СЏС‡РµР№РєРё None РёР»Рё РїСѓСЃС‚С‹Рµ СЃС‚СЂРѕРєРё.
    """
    for v in values:
        if v is None:
            continue
        if str(v).strip() != "":
            return False
    return True




def parse_product_column(header: str, value: Any) -> Optional[Dict]:
    """
    РџР°СЂСЃРёС‚ РєРѕР»РѕРЅРєСѓ СЃ С‚РѕРІР°СЂРѕРј.
    Р’РѕР·РІСЂР°С‰Р°РµС‚ СЃР»РѕРІР°СЂСЊ СЃ С‚РёРїРѕРј С‚РѕРІР°СЂР°, СЂР°Р·РјРµСЂРѕРј, С†РІРµС‚РѕРј Рё РєРѕР»РёС‡РµСЃС‚РІРѕРј
    """
    if not value or str(value).strip() == "":
        return None

    header_lower = header.lower()
    value_str = str(value).strip()

    # РћРїСЂРµРґРµР»СЏРµРј С‚РёРї С‚РѕРІР°СЂР°
    product_type = None
    for ptype, keywords in PRODUCT_KEYWORDS.items():
        if any(kw in header_lower for kw in keywords):
            product_type = ptype
            break

    if not product_type:
        return None

    result = {
        'type': product_type,
        'name': header[:200],  # РѕР±СЂРµР·Р°РµРј РґР»РёРЅРЅРѕРµ РЅР°Р·РІР°РЅРёРµ
        'size': None,
        'color': None,
        'quantity': 1  # РїРѕ СѓРјРѕР»С‡Р°РЅРёСЋ 1, РµСЃР»Рё РЅРµ СѓРєР°Р·Р°РЅРѕ РёРЅРѕРµ
    }

    # РџС‹С‚Р°РµРјСЃСЏ РёР·РІР»РµС‡СЊ СЂР°Р·РјРµСЂ
    size_match = re.search(r'СЂР°Р·РјРµСЂ\s+([\d\-SML\/]+)', header_lower, re.IGNORECASE)
    if size_match:
        result['size'] = size_match.group(1)

    # РџС‹С‚Р°РµРјСЃСЏ РёР·РІР»РµС‡СЊ С†РІРµС‚
    color_match = re.search(r'С†РІРµС‚\s+([Р°-СЏС‘]+)', header_lower, re.IGNORECASE)
    if color_match:
        result['color'] = color_match.group(1)

    # Р•СЃР»Рё Р·РЅР°С‡РµРЅРёРµ - С‡РёСЃР»Рѕ, СЌС‚Рѕ РјРѕР¶РµС‚ Р±С‹С‚СЊ РєРѕР»РёС‡РµСЃС‚РІРѕ
    if value_str.isdigit() and int(value_str) > 0:
        result['quantity'] = int(value_str)

    return result

def make_product_title(p: dict) -> str:
    """
    p: {type, size, color, name, quantity}
    """
    t = p.get("type")

    if t == "socks":
        base = "РќРѕСЃРєРё Р±РµРіРѕРІС‹Рµ"
    elif t == "headband":
        base = "РџРѕРІСЏР·РєР° СЃРїРѕСЂС‚РёРІРЅР°СЏ"
    elif t == "belt":
        base = "РџРѕСЏСЃ РґР»СЏ СЃРѕСЂРµРІРЅРѕРІР°РЅРёР№"
    elif t == "mug":
        base = "РљСЂСѓР¶РєР°"
    elif t == "donation":
        base = "Р”РѕРЅР°С‚ РІ РїСЂРёСЋС‚"
    elif t == "insurance":
        base = "РЎС‚СЂР°С…РѕРІРєР°"
    elif t == "sticker":
        base = "РќР°РєР»РµР№РєР°"
    else:
        base = "РўРѕРІР°СЂ"

    parts = [base]

    size = p.get("size")
    color = p.get("color")

    if size:
        parts.append(str(size))
    if color:
        parts.append(str(color))

    return ", ".join(parts)[:200]



def build_order_items(row: Dict[str, Any]) -> List[Dict]:
    """
    РЎРѕР±РёСЂР°РµС‚ РІСЃРµ С‚РѕРІР°СЂС‹ РёР· СЃС‚СЂРѕРєРё Excel
    """
    items = []

    for header, value in row.items():
        # РџСЂРѕРІРµСЂСЏРµРј, РїРѕС…РѕР¶Рµ Р»Рё РЅР° С‚РѕРІР°СЂРЅСѓСЋ РєРѕР»РѕРЅРєСѓ
        if any(kw in header.lower() for kw in ['РЅРѕСЃРѕРє', 'РЅРѕСЃРєРё', 'РїРѕРІСЏР·Рє', 'РїРѕСЏСЃ', 'РєСЂСѓР¶Рє', 'РґРѕРЅР°С‚']):
            product_data = parse_product_column(header, value)
            if product_data:
                items.append(product_data)

    return items


def get_or_create_product(product_data: Dict) -> Product:
    """
    РќР°С…РѕРґРёС‚ РёР»Рё СЃРѕР·РґР°РµС‚ С‚РѕРІР°СЂ РїРѕ РґР°РЅРЅС‹Рј РёР· Excel
    """
    # Р¤РѕСЂРјРёСЂСѓРµРј РЅР°Р·РІР°РЅРёРµ С‚РѕРІР°СЂР°
    name_parts = [product_data['name']]
    if product_data.get('size'):
        name_parts.append(f"СЂР°Р·РјРµСЂ {product_data['size']}")
    if product_data.get('color'):
        name_parts.append(f"С†РІРµС‚ {product_data['color']}")

    product_name = make_product_title(product_data)

    # РџС‹С‚Р°РµРјСЃСЏ РЅР°Р№С‚Рё СЃСѓС‰РµСЃС‚РІСѓСЋС‰РёР№
    product, created = Product.objects.get_or_create(
        name=product_name,
        defaults={
            'type': product_data['type'],
            'variant': product_data.get('color', ''),
            'size': product_data.get('size', ''),
        }
    )

    return product



class ExcelProcessor:
    @staticmethod
    @transaction.atomic
    def process_batch(batch: ImportBatch) -> dict:
        headers = []
        rows = []

        ext = os.path.splitext(batch.file.name)[1].lower()

        file_path = batch.file.path
        if ext == ".xls":
            from apps.imports.convert import convert_xls_to_xlsx
            file_path = convert_xls_to_xlsx(file_path)

        elif ext != ".xlsx":
            raise ValidationError("РџРѕРґРґРµСЂР¶РёРІР°РµС‚СЃСЏ С‚РѕР»СЊРєРѕ .xls РёР»Рё .xlsx")

        headers, rows = read_xlsx(file_path)


        for row_num, values in rows:
            data = row_to_dict(headers, values)

            raw = RawExcelRow.objects.create(
                batch=batch,
                row_number=row_num,
                raw_data=data,
            )



        if not batch.file:
            raise ValidationError("Р¤Р°Р№Р» РЅРµ Р·Р°РіСЂСѓР¶РµРЅ.")
        ext = os.path.splitext(batch.file.name)[1].lower()
        file_path = batch.file.path
        if ext == ".xls":
            from apps.imports.convert import convert_xls_to_xlsx
            file_path = convert_xls_to_xlsx(file_path)
        elif ext != ".xlsx":
            raise ValidationError("РџРѕРґРґРµСЂР¶РёРІР°РµС‚СЃСЏ С‚РѕР»СЊРєРѕ .xls РёР»Рё .xlsx")

        headers, rows = read_xlsx(file_path)


        RawExcelRow.objects.filter(batch=batch).delete()

        created_rows = 0
        created_clients = 0
        updated_clients = 0
        errors = 0
        skipped_empty = 0

        for row_num, values in rows:
            # 1) РїСЂРѕРїСѓСЃРєР°РµРј СЂРµР°Р»СЊРЅРѕ РїСѓСЃС‚С‹Рµ СЃС‚СЂРѕРєРё
            if not any(v is not None and _s(v) != "" for v in values):
                skipped_empty += 1
                continue

            data = row_to_dict(headers, values)

            raw = RawExcelRow.objects.create(
                batch=batch,
                row_number=row_num,
                raw_data=data,
            )
            created_rows += 1

            phone = normalize_phone(pick_any(data, exact=EXACT["phone"]))
            if not phone:
                raw.error_message = "РџСѓСЃС‚РѕР№/РЅРµРєРѕСЂСЂРµРєС‚РЅС‹Р№ С‚РµР»РµС„РѕРЅ"
                raw.save(update_fields=["error_message"])
                errors += 1
                continue

            city = _s(pick_any(data, exact=EXACT["city"]))
            contact = _s(pick_any(data, contains=CONTAINS["contact_text"]))
            pets = _s(pick_any(data, contains=CONTAINS["pets_text"]))
            email = _s(pick_any(data, exact=EXACT["email"]))
            address = build_address(data)
            full_name = build_full_name(data)
            notes = build_notes(data)



            dob = normalize_date(pick_any(data, exact=EXACT["dob"]))

            defaults = {
                "name": full_name,
                "email": email,
                "city": city,
                "address": address,  # в†ђ РёСЃРїСЂР°РІР»РµРЅРЅС‹Р№ Р°РґСЂРµСЃ
                "contact": contact,
                "pets": pets,
                "notes": notes,  # в†ђ РЅРѕРІС‹Рµ Р·Р°РјРµС‚РєРё
                "dob": dob or None,
            }

            if model_has_field(Client, "email"):
                defaults["email"] = email

            if model_has_field(Client, "city"):
                defaults["city"] = city

            if model_has_field(Client, "address"):
                defaults["address"] = address

            if model_has_field(Client, "contact"):
                defaults["contact"] = contact

            if model_has_field(Client, "pets"):
                defaults["pets"] = pets

            if model_has_field(Client, "dob"):
                defaults["dob"] = dob or None

            if model_has_field(Client, "notes"):
                defaults["notes"] = notes or None

            client, created = Client.objects.update_or_create(
                phone=phone,
                defaults=defaults,
            )

            raw.linked_client = client
            raw.save(update_fields=["linked_client"])

            if created:
                created_clients += 1
            else:
                updated_clients += 1


            tech_loc = get_tech_stock_location()
            distance = _s(pick_any(data, exact=EXACT["distance"]))
            bib = _s(pick_any(data, exact=EXACT["bib_number"]))
            chip = _s(pick_any(data, exact=EXACT["chip_number"]))
            product = make_product_title(data)


            order = Order.objects.create(
                client=client,
                event=batch.event,
                distance_text=distance,
                status='new',
                payment_status='NOT_PAID',
                payment_type='cash',
                registration_date=timezone.now(),
                comments=f"РќРѕРјРµСЂ: {bib}, Р§РёРї: {chip}".strip(", "),
                stock_location=tech_loc,

            )


            items_data = build_order_items(data)
            for item_data in items_data:
                product = get_or_create_product(item_data)

                OrderItem.objects.create(
                    order=order,
                    product=product,
                    quantity=item_data['quantity'],
                    price=0,
                )


            raw.linked_client = client
            raw.linked_order = order
            raw.save(update_fields=["linked_client", "linked_order"])

            if created:
                created_clients += 1
            else:
                updated_clients += 1


        return {
            "rows_saved": created_rows,
            "clients_created": created_clients,
            "clients_updated": updated_clients,
            "orders_created": created_rows - errors,
            "errors": errors,
            "skipped_empty_rows": skipped_empty,
        }


